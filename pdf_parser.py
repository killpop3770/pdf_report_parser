import camelot
import json
import os
from openpyxl import load_workbook
import time

startTime = time.time()

directory = "reports"
absDirectory = os.path.abspath(directory)
files = os.listdir(absDirectory)

wb = load_workbook('./first.xlsx')
sheet = wb['Лист1']

mode = 'a' if os.path.exists('storage_log.txt') else 'w'
file = open('storage_log.txt', mode)

for file in files:
	os.chdir(absDirectory)
	
	if "ЦО" in file or "ХВС" in file:
		continue

	tables = camelot.read_pdf(file)
	os.chdir("..")
	tables.export('foo.json', f='json', compress=False)

	string = " ".join(file.split("_")[:-4])
	string = string.replace("д. ", "д.")
	string = str(string.replace("к. ", "корп.")[5:])
	print("File name: "+string)

	with open('storage_log.txt', 'tr') as f:
		for line in f:
			if line == string+"\n":
				continue
			else:
				for nameCell in range(3, sheet.max_row):
					address = str(sheet[f'B{nameCell}'].value)
					if string == address:

							with open('foo-page-1-table-1.json') as report:
								data = json.loads(report.read()) 
								count = 0

								try:
									for i in range(2, len(data)-1):
										T = data[i].get('5')
										if T != "" or T is not None:
												Tflow = T.replace(',', '.')
												if float(Tflow) < 60.0:
													count+=1
										else:
											Tflow = 60.0
								except Exception as e:
									print("Bad data\n")
								finally:
									Tflow = 60.0

								position = str(nameCell)

								hours = sheet[f"D{position}"]
								Tflow = sheet[f"E{position}"]
								days = sheet[f"F{position}"]

								hours.value = data[-1].get('10')
								Tflow.value = data[-1].get('5')
								days.value = count

								wb.save("first.xlsx")

	with open('storage_log.txt', 'ta') as f:
		f.write(f"{string}\n")

	print("RunTime: "+str(time.time()-startTime))